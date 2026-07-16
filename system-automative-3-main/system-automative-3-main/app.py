import streamlit as st
import subprocess
import os
import sys
import pandas as pd
import json
import re
import io
import shlex
from typing import List
from streamlit_option_menu import option_menu
import base64
from pathlib import Path
from docx import Document
from docx.shared import Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.oxml.ns import qn
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent

DASHBOARD_DIR = BASE_DIR / "src" / "dashboard_0210"

if str(DASHBOARD_DIR) not in sys.path:
    sys.path.insert(0, str(DASHBOARD_DIR))

REPORTS_DIR = DASHBOARD_DIR / "reports"
HISTORY_DIR = DASHBOARD_DIR / "history"
IMAGES_DIR = DASHBOARD_DIR / "images"
SCRIPTS_DIR = DASHBOARD_DIR / "scripts"
CURRENT_DIR = DASHBOARD_DIR
TEMPLATES_DIR = DASHBOARD_DIR / "templates"
NUCLEI_BIN = str(Path.home() / "go" / "bin" / "nuclei")
NUCLEI_TEMPLATES_DIR = DASHBOARD_DIR / "nuclei-templates"
GUIDE_DIR = BASE_DIR / "src" / "guides"
ANSIBLE_CONFIG_PATH = CURRENT_DIR / "ansible.cfg"


def build_ansible_env() -> dict:
    env = os.environ.copy()
    env["ANSIBLE_CONFIG"] = str(ANSIBLE_CONFIG_PATH)
    return env

def cleanup_reports():
    import shutil
    report_dir = CURRENT_DIR / "reports"
    if report_dir.exists():
        for f in report_dir.glob("*_result.txt"):
            try:
                f.unlink()
            except:
                pass

def load_template(name: str) -> str:
    path = TEMPLATES_DIR / name
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8")

def execute_nuclei_command(command_text: str):
    command_text = (command_text or "").strip()
    if not command_text:
        return None, "명령어를 입력해주세요."

    try:
        cmd = shlex.split(command_text)
    except ValueError as e:
        return None, f"명령어 파싱 오류: {e}"

    if not cmd or cmd[0] != "nuclei":
        return None, "보안을 위해 nuclei 명령어만 실행할 수 있습니다."
    include_args: List[str] = []
    for i, token in enumerate(cmd[:-1]):
        if token in {"-t", "-templates"}:
            tpl_value = cmd[i + 1]
            if "/" in tpl_value and "-it" not in cmd and "-include-templates" not in cmd:
                include_args.extend(["-it", tpl_value])
    if include_args:
        cmd.extend(include_args)

    if "-j" not in cmd and "-jsonl" not in cmd:
        cmd.append("-j")
    if "-silent" not in cmd:
        cmd.append("-silent")

    try:
        proc = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=3600,
            cwd=str(CURRENT_DIR),
        )
    except subprocess.TimeoutExpired:
        return None, "명령어 실행 시간이 1시간을 초과했습니다."
    except Exception as e:
        return None, f"nuclei 실행 중 오류: {e}"

    parsed_json = []
    non_json_lines = []
    for raw_line in (proc.stdout or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            parsed_json.append(json.loads(line))
        except json.JSONDecodeError:
            non_json_lines.append(line)

    return {
        "cmd": cmd,
        "returncode": proc.returncode,
        "stdout": proc.stdout or "",
        "stderr": proc.stderr or "",
        "json": parsed_json,
        "non_json_lines": non_json_lines,
        #"include_args_added": include_args, 
    }, None

def normalize_auto_target(raw_target: str):
    target = (raw_target or "").strip()
    if not target:
        return "", ""

    # Allow convenient SSH-style input like: "ssh user@host"
    if target.lower().startswith("ssh "):
        target = target[4:].strip()

    if "@" in target and not target.startswith(("http://", "https://")):
        target = target.split("@", 1)[1].strip()

    return target, (raw_target or "").strip()

def summarize_nuclei_error(stderr_text: str, returncode: int) -> str:
    if not (stderr_text or "").strip():
        if returncode == 0:
            return "실행 성공(탐지 결과 없음)"
        return f"nuclei 실행 실패 (rc={returncode})"

    cleaned = re.sub(r"\x1b\[[0-9;]*m", "", stderr_text)
    lines = [line.strip() for line in cleaned.splitlines() if line.strip()]
    if not lines:
        if returncode == 0:
            return "실행 성공(탐지 결과 없음)"
        return f"nuclei 실행 실패 (rc={returncode})"

    for line in reversed(lines):
        lower = line.lower()
        if "could not run nuclei:" in lower:
            return line.split(":", 1)[1].strip() if ":" in line else line
        if "error" in lower or "failed" in lower or "no templates" in lower:
            return line
    if returncode == 0:
        return "실행 성공(탐지 결과 없음)"
    return lines[-1]

def load_image_base64(path: Path) -> str:
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except FileNotFoundError:
        return ""

RAPA_LOGO = load_image_base64(IMAGES_DIR / "rapa.png")
AUTOEVER_LOGO = load_image_base64(IMAGES_DIR / "hyundai_autoever.jpg")

st.set_page_config(
    page_title="Linux Security Dashboard",
    layout="wide",
    initial_sidebar_state="collapsed"
)

if "page" not in st.session_state:
    st.session_state.page = "main"

st.markdown("""
<style>
html, body {
    height: 100%;
}

.block-container {
    display: flex;
    flex-direction: column;

    padding-top: 0;
    padding-left: 0;
    padding-right: 0;
    padding-bottom: 0 !important;
    margin-bottom: 0 !important;
}

.hero-wrapper {
    width: 100%;
    margin-left: 0;
}

.hero {
    position: relative;
    width: 100%;
    min-height: 95vh;

    display: flex;
    align-items: center;
    justify-content: center;

    background:
        linear-gradient(
            to right,
            rgba(0,0,0,0.55) 0%,
            rgba(0,0,0,0.65) 40%,
            rgba(0,0,0,0.75) 100%
        ),
        url("https://images.unsplash.com/photo-1558494949-ef010cbdcc31");

    background-size: cover;
    background-position: center;
    background-repeat: no-repeat;

    background-attachment: fixed;

    transition: min-height 0.5s ease;
}

.hero-content {
    position: relative;
    z-index: 2;
    max-width: 1000px;
    text-align: center;
    color: #ffffff;
    padding: 0 24px;
}

.hero-content h1 {
    font-size: clamp(44px, 4.5vw, 72px);
    font-weight: 700;
    letter-spacing: -1px;
    margin-bottom: 16px;
}

.hero-content p {
    font-size: clamp(18px, 1.3vw, 24px);
    opacity: 0.9;
    line-height: 1.7;
}

.hero.shrink {
    min-height: 240px;
}

.hero.shrink .hero-content h1 {
    font-size: 32px;
}

.hero.sidebar-open .hero-content {
    transform: translate(calc(-50% + 160px), -50%);
}

.section {
    max-width: 1100px;
    margin: auto;
    padding: 80px 20px 120px;
}

section[data-testid="stSidebar"] {
    background-color: #f2f2f2;
}

.nav-link {
    margin: 6px 8px;
    padding: 10px 14px !important;

    font-size: 16px;
    color: #333 !important;
    border-radius: 14px !important;
}

.nav-link:hover {
    background-color: #e5e5e5 !important;
}

.nav-link.active,
.nav-link-selected {
    background-color: #dcdcdc !important;
    color: #000 !important;
    font-weight: 700 !important;
}

.nav-link i {
    font-size: 18px;
}

button[data-testid="collapsedControl"] {
    display: flex !important;
    align-items: center;
    gap: 6px;

    padding: 6px 12px !important;
    border-radius: 20px;

    background-color: #f2f2f2;
    color: #444;
    font-weight: 600;
}

button[data-testid="collapsedControl"]::after {
    content: "menu";
    font-size: 14px;
    letter-spacing: 0.5px;
}

button[data-testid="collapsedControl"]:hover {
    background-color: #e0e0e0;
}

body {
    background-color: #f7f9fc;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
}

.hero-cta {
    margin-top: 30px;
    display: flex;
    gap: 16px;
    justify-content: center;
}

.cta-primary {
    background: #ffffff;
    color: #0b1220;
    padding: 12px 22px;
    border-radius: 10px;
    font-weight: 700;
    text-decoration: none;
    transition: all .2s ease;
}

.cta-primary:hover {
    transform: translateY(-3px);
    box-shadow: 0 16px 36px rgba(0,0,0,0.2);
}

.cta-outline {
    border: 1px solid rgba(255,255,255,0.6);
    color: #ffffff;
    padding: 12px 22px;
    border-radius: 10px;
    font-weight: 600;
    text-decoration: none;
    transition: all .2s ease;
}

.cta-outline:hover {
    background: rgba(255,255,255,0.1);
}

.section-title {
    font-size: 38px;
    font-weight: 700;
    letter-spacing: -0.5px;
    margin-bottom: 28px;
    color: #1f2937;
    text-align: center;
}

.section-subtitle {
    font-size: 18px;
    line-height: 1.9;
    color: #4b5563;
    max-width: 900px;
    text-align: center;
    margin: 0 auto;
}

.section-subtitle strong {
    color: #111827;
    font-weight: 600;
}

.section-subtitle a {
    color: #005BAC;
    font-weight: 600;
    text-decoration: none;
}

.section-subtitle a:hover {
    text-decoration: underline;
}

.kpi-strip {
    display: flex;
    justify-content: space-between;
    text-align: center;
    margin-bottom: 80px;
}

.kpi-box h3 {
    font-size: 34px;
    font-weight: 700;
    margin-bottom: 6px;
}

.kpi-box p {
    font-size: 14px;
    opacity: 0.6;
}

.feature-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 30px;
}

.feature-card {
    background: #ffffff;
    border-radius: 16px;
    padding: 32px;
    box-shadow: 0 12px 32px rgba(0,0,0,0.06);
    transition: all .25s ease;
}

.feature-card:hover {
    transform: translateY(-8px);
    box-shadow: 0 20px 48px rgba(0,0,0,0.1);
}

.feature-card h4 {
    font-size: 18px;
    margin-bottom: 12px;
    font-weight: 700;
    position: relative;
    padding-left: 14px;
}

.feature-card h4::before {
    content: "";
    position: absolute;
    left: 0;
    top: 4px;
    width: 4px;
    height: 18px;
    background: #2563eb;
    border-radius: 4px;
}

.feature-card p {
    font-size: 15px;
    opacity: 0.7;
    line-height: 1.6;
}

@media (max-width: 1200px) {
    .feature-grid {
        grid-template-columns: repeat(2, 1fr);
    }
}
@media (max-width: 640px) {
    .feature-grid {
        grid-template-columns: 1fr;
    }
}

.kpi-strip {
    display: flex;
    justify-content: space-between;
    text-align: center;
    margin-bottom: 80px;
    border-top: 1px solid #e5e7eb;
    border-bottom: 1px solid #e5e7eb;
    padding: 40px 0;
}

.kpi-box:not(:last-child) {
    border-right: 1px solid #e5e7eb;
}
</style>

<style>
section[data-testid="stAppViewContainer"] {
    padding-bottom: 0 !important;
    min-height: 100vh;
    display: flex;
    flex-direction: column;
}

section[data-testid="stAppViewContainer"] > .block-container {
    flex: 1;
}

section[data-testid="stMain"] {
    padding-bottom: 0 !important;
}
</style>

<style>
.diagnosis-wrapper {
    display: flex;
    justify-content: center;
    margin-top: 0 !important;
}

.diagnosis-card {
    width: 100%;
    max-width: 720px;
    background-color: #f8f9fa;
    padding: 32px 36px;
    border-radius: 18px;
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.05);
}

.diagnosis-title {
    text-align: center;
    font-weight: 700;
    margin-bottom: 8px;
}

.diagnosis-desc {
    text-align: center;
    opacity: 0.8;
    line-height: 1.6;
    margin-bottom: 28px;
}
</style>

<style>
.result-wrapper {
    max-width: 1200px;
    margin: 0 auto;
}

.result-wrapper [data-testid="stStatus"],
.result-wrapper [data-testid="stAlert"] {
    width: 100% !important;
    max-width: 100% !important;
}
.about-wrapper {
    margin-top: 120px;
    padding: 80px 0;
}

.about-header h2 {
    font-size: 32px;
    margin-bottom: 8px;
}

.about-header p {
    color: #777;
    margin-bottom: 60px;
}

.team-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
    gap: 40px;
}

.team-card {
    padding: 28px 0;
    border-top: 2px solid #111;
    transition: all 0.3s ease;
}

.team-card:hover {
    transform: translateY(-4px);
}

.team-name {
    font-size: 18px;
    font-weight: 600;
    margin-bottom: 6px;
}

.team-role {
    font-size: 14px;
    color: #555;
    margin-bottom: 10px;
}

.team-email {
    font-size: 14px;
    color: #888;
}
</style>

<style>
div[data-testid="stDownloadButton"] button {
    all: unset;
    cursor: pointer;
    color: #2563eb;
    font-size: 15px;
}

div[data-testid="stDownloadButton"] button:hover {
    text-decoration: underline;
}
</style>

<script>
const updateHeroState = () => {
    const hero = document.querySelector(".hero");
    if (!hero) return;

    const sidebar = document.querySelector(
        'section[data-testid="stSidebar"]'
    );
    const sidebarOpen = sidebar && sidebar.offsetWidth > 100;

    if (window.scrollY > 160 || sidebarOpen) {
        hero.classList.add("shrink");
    } else {
        hero.classList.remove("shrink");
    }

    if (sidebarOpen) {
        hero.classList.add("sidebar-open");
    } else {
        hero.classList.remove("sidebar-open");
    }
};

window.addEventListener("scroll", updateHeroState);

const observer = new MutationObserver(updateHeroState);
observer.observe(document.body, {
    attributes: true,
    childList: true,
    subtree: true
});

const resetHero = () => {
  const hero = document.querySelector(".hero");
  if (!hero) return;
  hero.classList.remove("shrink");
  hero.classList.remove("sidebar-open");
};

resetHero();

setTimeout(resetHero, 200);

setTimeout(updateHeroState, 300);
</script>
""", unsafe_allow_html=True)

# =========================================================
# Sidebar Navigation
# =========================================================
with st.sidebar:
    selected = option_menu(
        menu_title=None,
        options=["main", "진단 및 조치", "nuclei", "기록"],
        icons=["star-fill", "shield-check", "search", "clock-history"],
        menu_icon="list",
        default_index=0,
        styles={
            "container": {"padding": "8px"},
            "icon": {"font-size": "18px"},
            "nav-link": {
                "font-size": "16px",
                "margin": "6px",
                "border-radius": "14px",
            },
            "nav-link-selected": {
                "background-color": "#dcdcdc",
                "color": "#000",
            },
        }
    )
    page_map = {
        "main": "main",
        "진단 및 조치": "check",
        "nuclei": "nuclei",
        "기록": "history",
    }

    st.session_state.page = page_map[selected]

# =========================================================
# MAIN / 점검 페이지
# =========================================================
if st.session_state.page == "main":
    cleanup_reports()
    
    st.markdown("""
    <div class="hero-wrapper">
        <div class="hero">
            <div class="hero-content">
                <h1>Diagnosis And Remediation auto Engine</h1>
                <p>
                    by 치약좋지
                </p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
<div class="section" id="overview">
    <div class="section-title">
        Linux Vulnerability Diagnosis Automation Platform
    </div>
    <div style="
        width: 400px;
        height: 2px;
        background: rgba(0, 91, 172, 0.35);
        margin: 22px auto 40px auto;
        border-radius: 2px;
    "></div>
    <div class="section-subtitle">
        <br><br>본 플랫폼은 
        <a href="https://www.kisa.or.kr/2060204/form?postSeq=22&page=1"
        target="_blank">
        KISA 주요정보통신기반시설 기술적 취약점 분석 상세 가이드(2026)
        </a>
        를 기준으로 설계된 <br><strong>엔터프라이즈 리눅스 취약점 진단 자동화 플랫폼</strong>입니다.<br><br>
        수동 점검 중심의 비효율적인 운영 방식을 개선하고,
        표준화된 정책 기반 진단 체계를 자동화하여<br>
        조직 내 보안 수준을 일관되게 유지할 수 있도록 지원합니다.<br><br>
        <strong>단일 서버부터 대규모 인프라 환경까지 확장 가능한 보안 점검 서비스</strong>를 제공합니다.<br><br><br><br>
    </div>
        <div class="feature-grid" id="features">
            <div class="feature-card">
                <h4>Single Server Assessment</h4>
                <p>
                IP 입력 기반 실시간 취약점 자동 진단.<br>
                KISA 표준 항목 기반 정밀 점검 수행.
                </p>
            </div>
            <div class="feature-card">
                <h4>Bulk Server Inspection</h4>
                <p>
                CSV 업로드 기반 다수 서버 일괄 분석.<br>
                운영 환경에 최적화된 대규모 자동 점검 수행.
                </p>
            </div>
            <div class="feature-card">
                <h4>Automated Reporting</h4>
                <p>
                진단 결과 자동 정리 및 Word 보고서 생성.<br>
                감사 대응 및 문서화 지원.
                </p>
            </div>
            <div class="feature-card">
                <h4>CVE Intelligence Integration</h4>
                <p>
                설정 취약점 + 공개 취약점 동시 분석.<br>
                정책 기반 진단과 실시간 위협 인텔리전스 결합.
                </p>
            </div>
        </div>
        <div class="about-wrapper">
    <div class="about-header">
        <h2>About Us</h2>
    </div>
    <div class="team-grid">
        <div class="team-card">
            <div class="team-name">송연수</div>
            <div class="team-email">songyeonsu12@gmail.com</div>
        </div>
        <div class="team-card">
            <div class="team-name">김연진</div>
            <div class="team-email">kyj9750322@gmail.com</div>
        </div>
        <div class="team-card">
            <div class="team-name">김태훈</div>
            <div class="team-email">kevin9480@naver.com</div>
        </div>
        <div class="team-card">
            <div class="team-name">이희윤</div>
            <div class="team-email">youthgmldbs@gmail.com</div>
        </div>
        <div class="team-card">
            <div class="team-name">조수진</div>
            <div class="team-email">suujin1025@gmail.com</div>
        </div>
    </div>
</div>
</div>

    """, unsafe_allow_html=True)
    
elif st.session_state.page == "check":
    # ===============================
    # 배너
    # ===============================
    st.markdown("""
    <div style="
        width: 100%;
        overflow: hidden;
        box-shadow: 0 8px 24px rgba(0,0,0,0.08);
        margin-bottom: 32px;
    ">
        <img src="https://images.unsplash.com/photo-1550751827-4bd374c3f58b"
             style="width:100%; height:220px; object-fit:cover;">
    </div>
    """, unsafe_allow_html=True)

    # ===============================
    # 진단 페이지
    # ===============================
    st.markdown("""
    <div class="diagnosis-wrapper">
        <div class="diagnosis-card">
            <h3 class="diagnosis-title">⚙️ 취약점 진단</h3>
            <div class="diagnosis-desc">
            단일 서버에 대한 개별 진단과<br>
            다중 서버에 대한 일괄 진단을 지원합니다.<br>
            환경 규모에 따라 유연한 점검 방식을 선택할 수 있습니다.
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='height:80px'></div>", unsafe_allow_html=True)

    # ===============================
    # 사용자 입력
    # ===============================
    _, center, _ = st.columns([1, 3, 1])
    with center:
        tab1, tab2 = st.tabs(["🎯 개별 서버 진단", "📁 다중 서버 진단 (CSV)"])
        st.markdown("""
        <style>

        div[data-testid="stTabs"] button {
            font-size: 18px !important;
            font-weight: 700 !important;
            padding: 14px 28px !important;
            border-radius: 10px 10px 0 0 !important;
            border-bottom: none !important;
        }

        div[data-testid="stTabs"] div[role="tablist"]::after {
            display: none !important;
        }

        div[data-testid="stTabs"] button[aria-selected="true"] {
            background-color: #f2f2f2 !important;   /* 회색 배경 */
            color: #000 !important;
        }

        div[data-testid="stTabs"] button[aria-selected="false"] {
            background-color: transparent !important;
            color: #444 !important;
        }

        </style>
        """, unsafe_allow_html=True)

        st.markdown("""
        <style>
        .tooltip-container {
            position: relative;
            display: inline-block;
            cursor: pointer;
        }

        .tooltip-container .tooltip-text {
            visibility: hidden;
            width: 400px;
            background-color: #1e1e1e;
            color: #ffffff;
            text-align: left;
            border-radius: 12px;
            padding: 18px 20px;
            position: absolute;
            z-index: 999;

            top: 50%;
            left: 115%;
            transform: translateY(-50%) translateX(-10px);

            font-size: 15px;
            line-height: 1.7;
            opacity: 0;

            transition: all 0.35s ease;
            box-shadow: 0px 8px 22px rgba(0,0,0,0.35);
        }

        .tooltip-container:hover .tooltip-text {
            visibility: visible;
            opacity: 1;
            transform: translateY(-50%) translateX(0px);
        }
        </style>
        """, unsafe_allow_html=True)

        with tab1:
            target_ip = st.text_input("대상 서버 IP", placeholder="192.168.x.x", key="single_ip")
            ssh_user = st.text_input("SSH 계정", value="", key="single_user")
            ssh_pw = st.text_input("SSH 비밀번호", type="password", key="single_pw")
            uploaded_file = None

        with tab2:
            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
            uploaded_file = st.file_uploader("서버 목록 CSV 업로드 (필수: ip, user, pw)", type=["csv"], key="bulk_upload")
            if uploaded_file:
                try:
                    df_targets = pd.read_csv(uploaded_file, encoding='utf-8-sig')
                    st.dataframe(df_targets, use_container_width=True, height=150)
                except Exception as e:
                    st.error(f"CSV 읽기 실패: {e}")

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        start_btn = st.button("🚀 진단 시작", use_container_width=True)

        if start_btn:
            # 이전 진단 관련 상태 전부 초기화
            for key in [
                "latest_result_ip",
                "latest_result_df",
                "current_security_level",
                "before_security_level",
                "remedy_done",
                "remedy_codes",
                "before_remedy_df",
                "remedy_select"
            ]:
                st.session_state.pop(key, None)

    st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)
    st.divider()

    # ===============================
    # 진단 실행
    # ===============================
    _, result_center, _ = st.columns([0.3, 6, 0.3])

    if start_btn:
        inventory_path = CURRENT_DIR / "temp_inventory.ini"
        playbook_path = CURRENT_DIR / "check_playbook.yml"
        
        # 대상 확인 및 인벤토리 생성
        valid_target = False
        with open(inventory_path, "w", encoding="utf-8") as f:
            f.write("[targets]\n")
            
            # CSV 파일이 업로드된 경우 (탭2)
            if uploaded_file is not None:
                for _, row in df_targets.iterrows():
                    f.write(f"{row['ip']} ansible_user={row['user']} ansible_password={row['pw']} ansible_become_password={row['pw']}\n")
                display_msg = "다중 서버"
                valid_target = True
            
            # 개별 IP가 입력된 경우 (탭1)
            elif target_ip:
                f.write(f"{target_ip} ansible_user={ssh_user} ansible_password={ssh_pw} ansible_become_password={ssh_pw}\n")
                display_msg = target_ip
                valid_target = True

        if not valid_target:
            st.error("진단 대상을 입력하거나 CSV 파일을 업로드해주세요!")
        else:
            with result_center:
                with st.status(f"🌐 {display_msg} 진단 중...", expanded=True) as status:
                    result = subprocess.run(
                        ["ansible-playbook", "-i", str(inventory_path), str(playbook_path)],
                        capture_output=True,
                        text=True,
                        cwd=str(CURRENT_DIR),
                        env=build_ansible_env(),
                    )

                    if result.returncode == 0:
                        status.update(label="✅ 진단 완료!", state="complete")
                        # 단일 진단일 경우 바로 결과 세션 저장
                        if uploaded_file is None:
                            st.session_state["latest_result_ip"] = target_ip
                        st.balloons()
                        st.success(f"🎉 {display_msg} 점검 성공!")
                    else:
                        status.update(label="❌ 진단 실패", state="error")
                        st.error("진단 실행 중 오류가 발생했습니다.")
                        st.code(result.stderr)

                        # 에러나면 주석 풀고 디버깅용으로 사용하새요 ~
                        st.write("Return Code:", result.returncode)
                        st.write("STDOUT:")
                        st.code(result.stdout)

                        st.write("STDERR:")
                        st.code(result.stderr)

    # =====================================================
    # RESULT REPORT
    # =====================================================
    report_dir = CURRENT_DIR / "reports"

    if report_dir.exists():
        report_files = sorted([f.name for f in report_dir.glob("*_result.txt")])

        if report_files:
            _, result_center, _ = st.columns([0.3, 6, 0.3])
            with result_center:
                st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)
                st.markdown("### 📋 진단 결과 리포트 선택")

                selected_file = st.selectbox(
                    "결과를 확인할 서버를 선택하세요",
                    report_files,
                    index=0,
                    help="점검이 완료된 서버의 IP 목록입니다."
                )
                st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)
                recent_ip = selected_file.replace("_result.txt", "")
                st.session_state["latest_result_ip"] = recent_ip
                report_path = report_dir / selected_file

                st.markdown(
                    f"<h3>📊 {recent_ip} 진단 결과</h3>",
                    unsafe_allow_html=True
                )
                st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)

                try:
                    parsed_results = []
                    with open(report_path, "r", encoding="utf-8") as f:
                        for line in f:
                            line = line.strip()
                            if line.startswith("{") and line.endswith("}"):
                                data = json.loads(line)
                                parsed_results.append({
                                    "코드": data.get("code"),
                                    "중요도": data.get("severity"),
                                    "항목": data.get("item"),
                                    "상태": data.get("status"),
                                    "상세 사유": data.get("reason"),
                                })

                    if parsed_results:
                        df = pd.DataFrame(parsed_results)
                        df = df[["코드", "중요도", "항목", "상태", "상세 사유"]]

                        df = df[df["코드"].notna()]

                        df["STATUS_ORDER"] = df["상태"].apply(
                            lambda x: 0 if "취약" in str(x) else 1
                        )

                        df["U_NUM"] = df["코드"].str.extract(r'U-(\d+)').astype(int)

                        df = df.sort_values(
                            by=["STATUS_ORDER", "U_NUM"],
                            ascending=[True, True]
                        )

                        df = df.drop(columns=["STATUS_ORDER", "U_NUM"])
                        df = df.reset_index(drop=True)
                        
                        #---- 보안수준 계산 ----
                        score_map = {"상": 10, "중": 8, "하": 6}
                        TOTAL_SCORE = 598

                        vuln_df_all = df[df["상태"] == "취약"]
                        vuln_score_sum = vuln_df_all["중요도"].map(score_map).sum()

                        security_level = round(((TOTAL_SCORE - vuln_score_sum) / TOTAL_SCORE) * 100, 2)

                        st.session_state["current_security_level"] = security_level
                        
                        st.markdown(
                            f"""
                            <div class="tooltip-container" style="font-size:19px; font-weight:bold;">
                                🔐 보안 수준 : {security_level}%
                                <div class="tooltip-text">
                                    <b>보안수준 계산식</b><br>
                                    *598 = 모든 항목을 취약이라 가정했을 때의 점수 합<br>
                                    (598 - 취약항목 점수합) ÷ 598 × 100<br><br>
                                    • 상 : 10점<br>
                                    • 중 : 8점<br>
                                    • 하 : 6점
                                </div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )

                        st.session_state["latest_result_df"] = df

                        def highlight_vulnerable(row):
                            if "취약" in str(row["상태"]):
                                return ["background-color: #ffe6e1"] * len(row)
                            return [""] * len(row)

                        st.dataframe(
                            df.style
                                .apply(highlight_vulnerable, axis=1)
                                .map(lambda x: "color:red; font-weight:bold;" if "취약" in str(x) else "color:green;",
                                    subset=["상태"])
                                .map(lambda x: "color:red;" if x == "상" else "color:orange;",
                                    subset=["중요도"]),
                            use_container_width=True,
                            height=420
                        )

                        st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)
                        if st.button(f"📊 {recent_ip} 결과 Excel로 보관함 저장"):

                            HISTORY_DIR = CURRENT_DIR / "history"
                            HISTORY_DIR.mkdir(exist_ok=True)

                            date_str = datetime.now().strftime("%Y-%m-%d")
                            file_time = datetime.now().strftime("%Y-%m-%d_%H%M%S")

                            excel_path = HISTORY_DIR / f"{recent_ip}_{file_time}.xlsx"

                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Diagnosis Result"

                            ws.merge_cells("A1:E1")
                            ws["A1"] = f"{date_str} 취약점 점검 결과"
                            ws["A1"].font = Font(size=16, bold=True)
                            ws["A1"].alignment = Alignment(horizontal="center")

                            ws.merge_cells("A2:E2")
                            ws["A2"] = f"대상 서버 : {recent_ip}"
                            ws["A2"].font = Font(size=12, bold=True)
                            ws["A2"].alignment = Alignment(horizontal="center")

                            current_security = st.session_state.get("current_security_level", 0)
                            ws.merge_cells("A3:E3")
                            ws["A3"] = f"보안 수준 : {current_security}%"
                            ws["A3"].font = Font(size=12, bold=True)
                            ws["A3"].alignment = Alignment(horizontal="center")

                            start_row = 5

                            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                                for c_idx, value in enumerate(row, 1):
                                    ws.cell(row=r_idx, column=c_idx, value=value)

                            vuln_fill = PatternFill(start_color="FFE6E1", end_color="FFE6E1", fill_type="solid")
                            red_font = Font(color="FF0000", bold=True)
                            green_font = Font(color="008000")
                            orange_font = Font(color="FF8C00")

                            from openpyxl.styles import Border, Side

                            thin = Side(style="thin")
                            border = Border(left=thin, right=thin, top=thin, bottom=thin)

                            for row in ws.iter_rows(min_row=start_row+1, max_row=ws.max_row):
                                status_cell = row[3]
                                severity_cell = row[1]

                                for cell in row:
                                    cell.border = border

                                if status_cell.value == "취약":
                                    for cell in row:
                                        status_cell.font = red_font

                                elif status_cell.value == "양호":
                                    status_cell.font = green_font

                                if severity_cell.value == "상":
                                    severity_cell.font = red_font
                                elif severity_cell.value == "중":
                                    severity_cell.font = orange_font

                            from openpyxl.utils import get_column_letter

                            for col_idx in range(1, ws.max_column + 1):
                                ws.column_dimensions[get_column_letter(col_idx)].width = 22

                            wb.save(excel_path)

                            st.success(f"📁 {recent_ip} Excel 리포트가 저장되었습니다.")

                            with open(excel_path, "rb") as f:
                                st.download_button(
                                    label="⬇️ Excel 다운로드",
                                    data=f.read(),
                                    file_name=excel_path.name,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                    else:
                        st.info(f"{recent_ip} 서버의 상세 진단 결과가 비어있습니다.")

                except Exception as e:
                    st.error(f"리포트 처리 중 오류 발생: {e}")

                # ===============================
                # 조치
                # ===============================
                st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)
                st.divider()
                st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

                st.markdown("### 🛠 조치")

                MANUAL_ONLY_CODES = [
                    "U-01", "U-02", "U-07", "U-09", "U-23", "U-25", "U-28", "U-30", "U-33",
                    "U-40", "U-45", "U-47", "U-49", "U-50", "U-51",
                    "U-56", "U-61", "U-64", "U-66"
                ]

                df = st.session_state.get("latest_result_df")
                if df is None or df.empty:
                    st.info("조치 기능을 사용하려면 유효한 진단 결과를 먼저 불러와주세요.")
                    vuln_df = pd.DataFrame()
                    auto_remedy_df = pd.DataFrame()
                    manual_df = pd.DataFrame()
                else:
                    vuln_df = df[df["상태"] == "취약"]
                    auto_remedy_df = vuln_df[~vuln_df["코드"].isin(MANUAL_ONLY_CODES)]
                    manual_df = vuln_df[vuln_df["코드"].isin(MANUAL_ONLY_CODES)]

                selected_codes = []
                remedy_btn = False

                # -------------------------------
                # 자동 조치 영역
                # -------------------------------
                if not auto_remedy_df.empty:

                    selected_codes = st.multiselect(
                        "조치할 취약 항목을 선택하세요",
                        options=auto_remedy_df["코드"].tolist(),
                        key="remedy_select"
                    )

                    remedy_btn = st.button("🛠 선택 항목 조치 실행", key="remedy_button")

                else:
                    st.success("자동 조치 가능한 취약 항목이 없습니다.")

                # -------------------------------
                # 수동 조치 안내
                # -------------------------------
                if not manual_df.empty:
                    manual_list = ", ".join(manual_df["코드"].tolist())
                    st.info(
                        f"""
                        🔎 {manual_list} 항목에 대해서는 점검자의 판단에 따라 수동 조치를 실시하세요.  
                        (PDF 가이드 참고)
                        """
                    )

                    rocky_pdf = GUIDE_DIR / "수동_조치_가이드라인_rocky_linux.pdf"
                    ubuntu_pdf = GUIDE_DIR / "수동_조치_가이드라인_ubuntu.pdf"

                    col1, col2 = st.columns(2)

                    with col1:
                        if rocky_pdf.exists():
                            with open(rocky_pdf, "rb") as f:
                                st.download_button(
                                    label="📘 Rocky Linux 수동 조치 가이드 다운로드",
                                    data=f.read(),
                                    file_name="수동_조치_가이드라인_rocky_linux.pdf",
                                    mime="application/pdf"
                                )

                    with col2:
                        if ubuntu_pdf.exists():
                            with open(ubuntu_pdf, "rb") as f:
                                st.download_button(
                                    label="📘 Ubuntu 수동 조치 가이드 다운로드",
                                    data=f.read(),
                                    file_name="수동_조치_가이드라인_ubuntu.pdf",
                                    mime="application/pdf"
                                )

                # ----------------------------------------------------
                # 조치 버튼 클릭
                # ----------------------------------------------------
                if remedy_btn:

                    if not selected_codes:
                        st.warning("조치할 항목을 선택해주세요.")
                    else:

                        before_df = df[df["코드"].isin(selected_codes)].copy()
                        st.session_state["before_remedy_df"] = before_df

                        score_map = {"상": 10, "중": 8, "하": 6}
                        TOTAL_SCORE = 598

                        vuln_before = df[df["상태"] == "취약"]
                        before_score = vuln_before["중요도"].map(score_map).sum()
                        before_security = round(((TOTAL_SCORE - before_score) / TOTAL_SCORE) * 100, 2)

                        st.session_state["before_security_level"] = before_security

                        inventory_path = CURRENT_DIR / "temp_inventory.ini"
                        remedy_playbook_path = CURRENT_DIR / "remedy_playbook.yml"
                        check_playbook_path = CURRENT_DIR / "check_playbook.yml"

                        with st.status("선택 항목 조치 실행 중...", expanded=True) as status:

                            for code in selected_codes:

                                result = subprocess.run(
                                    [
                                        "ansible-playbook",
                                        "-i", str(inventory_path),
                                        str(remedy_playbook_path),
                                        "-e", f"vuln_code={code}",
                                        "--limit", recent_ip
                                    ],
                                    capture_output=True,
                                    text=True,
                                    cwd=str(CURRENT_DIR),
                                    env=build_ansible_env(),
                                )

                                if result.returncode != 0:
                                    st.error(f"{code} 조치 실패")
                                    if (result.stdout or "").strip():
                                        st.caption("STDOUT")
                                        st.code(result.stdout)
                                    if (result.stderr or "").strip():
                                        st.caption("STDERR")
                                    st.code(result.stderr)
                                    status.update(label="❌ 조치 실패", state="error")
                                    break

                            else:
                                status.update(label="✅ 조치 완료", state="complete")

                                status.update(label="🔄 재진단 실행 중...", state="running")
                                
                                # 재진단
                                recheck = subprocess.run(
                                    [
                                        "ansible-playbook",
                                        "-i", str(inventory_path),
                                        str(check_playbook_path),
                                        "--limit", recent_ip
                                    ],
                                    capture_output=True,
                                    text=True,
                                    cwd=str(CURRENT_DIR),
                                    env=build_ansible_env(),
                                )

                                if recheck.returncode == 0:
                                    status.update(label="✅ 조치 및 재진단 완료", state="complete")
                                    st.session_state["remedy_done"] = True
                                    st.session_state["remedy_codes"] = selected_codes
                                    st.rerun()
                                else:
                                    status.update(label="❌ 재진단 실패", state="error")
                                    st.error("재진단 실패")
                                    st.code(recheck.stderr)

                # ----------------------------------------------------
                # 조치 결과 출력
                # ----------------------------------------------------
                if st.session_state.get("remedy_done") and df is not None and not df.empty:
                    st.success("✅ 조치 및 재진단 완료. 전체 결과를 확인하려면 위의 표를 다시 확인하세요.")

                    remedy_codes = st.session_state.get("remedy_codes", [])
                    before_df = st.session_state.get("before_remedy_df")

                    after_df = df[df["코드"].isin(remedy_codes)].copy()

                    compare_df = before_df.merge(
                        after_df,
                        on="코드",
                        suffixes=("_Before", "_After")
                    )

                    #-----보안수준계산
                    score_map = {"상": 10, "중": 8, "하": 6}
                    TOTAL_SCORE = 598

                    vuln_after = df[df["상태"] == "취약"]
                    after_score = vuln_after["중요도"].map(score_map).sum()
                    after_security = round(((TOTAL_SCORE - after_score) / TOTAL_SCORE) * 100, 2)

                    before_security = st.session_state.get("before_security_level", 0)

                    def get_result(row):
                        if row["상태_Before"] == "취약" and row["상태_After"] == "양호":
                            return "개선됨"
                        elif row["상태_After"] == "취약":
                            return "취약 유지"
                        else:
                            return "양호 유지"

                    compare_df["결과"] = compare_df.apply(get_result, axis=1)

                    compare_df = compare_df[[
                        "코드",
                        "결과",
                        "상태_Before",
                        "상세 사유_Before",
                        "상태_After",
                        "상세 사유_After",
                    ]]
                        
                    st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)
                    st.markdown("#### 📊 조치 전후 비교 결과표")
                    st.markdown(
                        f"""
                        <div style="font-size:18px; font-weight:600; margin-bottom:10px;">
                            🔐 보안 수준 : {before_security}% → {after_security}%
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                    def highlight_status(val):
                        if val == "양호":
                            return "color:green; font-weight:bold;"
                        elif val == "취약":
                            return "color:red; font-weight:bold;"
                        return ""

                    def highlight_result(val):
                        if val == "개선됨":
                            return "background-color:#d4edda; color:#155724; font-weight:bold;"
                        elif val == "취약 유지":
                            return "background-color:#f8d7da; color:#721c24; font-weight:bold;"
                        return ""

                    styled_df = compare_df.style \
                        .map(highlight_status, subset=["상태_Before", "상태_After"]) \
                        .map(highlight_result, subset=["결과"])

                    st.dataframe(styled_df, use_container_width=True)

                    del st.session_state["remedy_done"]
                    del st.session_state["remedy_codes"]
                    del st.session_state["before_remedy_df"]

                    if "remedy_select" in st.session_state:
                        del st.session_state["remedy_select"]
                elif st.session_state.get("remedy_done"):
                    st.warning("재진단 결과를 불러오지 못해 조치 전후 비교를 표시할 수 없습니다.")

            st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)



# =========================================================
# NUCLEI PAGE
# =========================================================
elif st.session_state.page == "nuclei":

    # ===============================
    # 배너 (check 페이지와 동일)
    # ===============================
    st.markdown("""
        <div style="
            width: 100%;
            overflow: hidden;
            box-shadow: 0 8px 24px rgba(0,0,0,0.08);
            margin-bottom: 32px;
        ">
            <img src="https://images.unsplash.com/photo-1550751827-4bd374c3f58b"
                style="width:100%; height:220px; object-fit:cover;">
        </div>
        """, unsafe_allow_html=True)

    # ===============================
    # 카드 헤더
    # ===============================
    st.markdown("""
    <div class="diagnosis-wrapper">
        <div class="diagnosis-card">
            <h3 class="diagnosis-title">🧪 Nuclei 스캔</h3>
            <div class="diagnosis-desc">
                Nuclei 템플릿 기반 자동 스캔 및 명령어 실행을 지원합니다.
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='height:80px'></div>", unsafe_allow_html=True)    

    # ===============================
    # 중앙 정렬
    # ===============================
    _, center, _ = st.columns([1, 3, 1])
    with center:
        tab1, tab2 = st.tabs(["🚀 자동 스캔", "🖥 명령어 직접 실행"])
        
        st.markdown("""
        <style>

        div[data-testid="stTabs"] button {
            font-size: 18px !important;
            font-weight: 700 !important;
            padding: 14px 28px !important;
            border-radius: 10px 10px 0 0 !important;
            border-bottom: none !important;
        }

        div[data-testid="stTabs"] div[role="tablist"]::after {
            display: none !important;
        }

        div[data-testid="stTabs"] button[aria-selected="true"] {
            background-color: #f2f2f2 !important;   /* 회색 배경 */
            color: #000 !important;
        }

        div[data-testid="stTabs"] button[aria-selected="false"] {
            background-color: transparent !important;
            color: #444 !important;
        }

        </style>
        """, unsafe_allow_html=True)


        with tab1:

            st.markdown("""
            <div style="
                background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
                padding: 24px 28px;
                border-radius: 18px;
                margin-top: 20px;
                margin-bottom: 30px;
                border: 1px solid #e2e8f0;
                box-shadow: 0 4px 12px rgba(0,0,0,0.04);
            ">
                <!-- 제목 영역 -->
                <div style="
                    display: grid;
                    grid-template-columns: 28px 1fr;
                    align-items: center;
                    margin-bottom: 14px;
                ">
                    <div style="font-size:18px;">🔎</div>
                    <div style="font-weight:700; font-size:17px;">
                        자동 스캔 모드 안내
                    </div>
                </div>
                <!-- 내용 영역 -->
                <div style="
                    display: grid;
                    grid-template-columns: 165px 1fr;
                    row-gap: 10px;
                    column-gap: 12px;
                    font-size: 14.5px;
                    line-height: 1.55;
                ">
                    <div style="font-weight:600; color:#1e293b;">웹 기본 스캔</div>
                    <div style="color:#334155;">웹 CVE 및 보안 설정 오류 점검</div>
                    <div style="font-weight:600; color:#1e293b;">웹 확장 스캔</div>
                    <div style="color:#334155;">취약점, 정보 노출, 기본 계정, Takeover 탐지</div>
                    <div style="font-weight:600; color:#1e293b;">네트워크 스캔</div>
                    <div style="color:#334155;">포트/서비스 기반 취약점 점검</div>
                    <div style="font-weight:600; color:#1e293b;">DNS / SSL 스캔</div>
                    <div style="color:#334155;">인증서 및 DNS 설정 점검</div>
                    <div style="font-weight:600; color:#1e293b;">DAST 스캔</div>
                    <div style="color:#334155;">실행 중인 웹 애플리케이션 대상 동적 취약점 점검</div>
                    <div style="font-weight:600; color:#1e293b;">전체 템플릿 스캔</div>
                    <div style="color:#334155;">모든 템플릿 기반 광범위 검사 (시간 소요 ↑)</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            auto_target = st.text_input(
                "스캔 대상",
                placeholder="예) ssh song@192.168.xxx.xxx",
                help="ssh user@host 형식도 입력 가능하며 내부적으로 host/IP로 변환합니다. 웹/DAST 스캔은 URL(예: https://example.com)을 사용하세요.",
                key="nuclei_auto_target",
            )
            auto_mode = st.selectbox(
                "스캔 모드",
                [   
                    "스캔 모드를 선택하세요",
                    "웹 기본 스캔",
                    "웹 확장 스캔",
                    "네트워크 스캔",
                    "DNS/SSL 스캔",
                    # "Linux 로컬 감사 스캔",
                    # "코드 취약점 스캔",
                    "DAST 스캔",
                    "전체 템플릿 스캔",
                ],
                index=0,
                key="nuclei_auto_mode",
            )
            sev_list = st.multiselect(
                "중요도 필터",
                options=["critical", "high", "medium", "low", "info"],
                key="nuclei_auto_severity",
            )

            # if st.button("🚀 자동 스캔 실행", use_container_width=True, key="nuclei_auto_run"):
            auto_run = st.button(
                "🚀 자동 스캔 실행",
                use_container_width=True,
                key="nuclei_auto_run"
            )

            if auto_run:
                if not auto_target.strip():
                    st.error("스캔 대상을 입력해주세요.")

                if auto_mode == "스캔 모드를 선택하세요":
                    st.warning("스캔 모드를 선택해주세요.")

                else:
                    severity_arg = ",".join(sev_list) if sev_list else "critical,high,medium,low,info"
                    normalized_target, original_target = normalize_auto_target(auto_target)
                    if not normalized_target:
                        st.error("스캔 대상을 올바르게 입력해주세요.")
                        st.stop()
                    target_q = shlex.quote(auto_target.strip())
                    http_cves = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "http" / "cves"))
                    http_mis = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "http" / "misconfiguration"))
                    http_vuln = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "http" / "vulnerabilities"))
                    http_exposures = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "http" / "exposures"))
                    http_default_logins = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "http" / "default-logins"))
                    http_takeovers = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "http" / "takeovers"))

                    network_cves = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "network" / "cves"))
                    network_exposures = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "network" / "exposures"))
                    network_vuln = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "network" / "vulnerabilities"))

                    dns_templates = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "dns"))
                    ssl_templates = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "ssl"))

                    dast_cves = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "dast" / "cves"))
                    dast_vuln = shlex.quote(str(NUCLEI_TEMPLATES_DIR / "dast" / "vulnerabilities"))
                    templates_root = shlex.quote(str(NUCLEI_TEMPLATES_DIR))

                    if auto_mode == "웹 기본 스캔":
                        if not normalized_target.startswith(("http://", "https://")):
                            st.error("웹 스캔은 URL 대상만 지원합니다. 예: https://example.com")
                            st.stop()
                        auto_cmd = (
                            f"nuclei -u {target_q} -t {http_cves} -t {http_mis} "
                            f"-severity {severity_arg} -rate-limit 50 -timeout 10"
                        )
                    elif auto_mode == "웹 확장 스캔":
                        if not normalized_target.startswith(("http://", "https://")):
                            st.error("웹 스캔은 URL 대상만 지원합니다. 예: https://example.com")
                            st.stop()
                        auto_cmd = (
                            f"nuclei -u {target_q} -t {http_vuln} -t {http_exposures} "
                            f"-t {http_default_logins} -t {http_takeovers} "
                            f"-severity {severity_arg} -rate-limit 50 -timeout 10"
                        )
                    elif auto_mode == "네트워크 스캔":
                        auto_cmd = (
                            f"nuclei -target {target_q} -t {network_cves} -t {network_exposures} "
                            f"-t {network_vuln} -severity {severity_arg} -rate-limit 50 -timeout 10"
                        )
                    elif auto_mode == "DNS/SSL 스캔":
                        auto_cmd = (
                            f"nuclei -target {target_q} -t {dns_templates} -t {ssl_templates} "
                            f"-severity {severity_arg} -rate-limit 50 -timeout 10"
                        )
                    # elif auto_mode == "Linux 로컬 감사 스캔":
                    #     auto_cmd = (
                    #         f"nuclei -target {target_q} -t {linux_audit} "
                    #         f"-code -esc -severity {severity_arg} -timeout 10"
                    #     )
                    # elif auto_mode == "코드 취약점 스캔":
                    #     auto_cmd = (
                    #         f"nuclei -target {target_q} -t {code_cves} -t {code_misconfig} "
                    #         f"-code -esc -severity {severity_arg} -timeout 10"
                    #     )
                    elif auto_mode == "DAST 스캔":
                        if not normalized_target.startswith(("http://", "https://")):
                            st.error("DAST 스캔은 URL 대상만 지원합니다. 예: https://example.com")
                            st.stop()
                        auto_cmd = (
                            f"nuclei -u {target_q} -t {dast_cves} -t {dast_vuln} "
                            f"-dast -severity {severity_arg} -rate-limit 50 -timeout 10"
                        )
                    else:
                        if not normalized_target.startswith(("http://", "https://")):
                            st.error("전체 템플릿 스캔은 URL 대상만 지원합니다. 예: https://example.com")
                        auto_cmd = (
                            f"nuclei -u {target_q} -t {templates_root} "
                            f"-severity {severity_arg} -rate-limit 50 -timeout 10"
                        )
                    if original_target != normalized_target:
                        st.info(f"입력 대상 정규화: `{original_target}` → `{normalized_target}`")

                    with st.status("Nuclei 자동 스캔 실행 중...", expanded=False):
                        result, err = execute_nuclei_command(auto_cmd)
                    if err:
                        st.error(err)
                    else:
                        st.session_state["nuclei_last_result"] = result
                        if result.get("returncode", 1) == 0:
                            st.success("Nuclei 자동 스캔이 완료되었습니다.")
                        else:
                            st.warning(f"Nuclei 실행은 끝났지만 오류 코드가 반환되었습니다. (rc={result.get('returncode')})")
                        # st.info("아래 `실행 결과` 섹션에서 JSON/로그를 확인하세요.")
            result = st.session_state.get("nuclei_last_result")
            if result:
                st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)
                st.markdown("#### 📊 실행 결과")
                # 디버깅용 
                # st.write(f"Return code: `{result['returncode']}`")
                # st.code(
                #     " ".join(shlex.quote(token) for token in result["cmd"]),
                #     language="bash"
                # )
                if result.get("include_args_added"):
                    st.caption(f"템플릿 로딩 안정화를 위해 자동 보정 인자 추가: `{result['include_args_added']}`")               
                
                st.caption(
                    f"JSON 건수: `{len(result.get('json', []))}` | "
                    f"STDOUT 라인: `{len((result.get('stdout') or '').splitlines())}` | "
                    f"STDERR 라인: `{len((result.get('stderr') or '').splitlines())}`"
                )
                st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

                no_findings = (
                    result.get("returncode", 1) == 0
                    and len(result.get("json", [])) == 0
                    and not (result.get("stdout") or "").strip()
                    and not (result.get("stderr") or "").strip()
                )
                if no_findings:
                    st.success("탐지된 취약점이 없습니다. (No findings)")

                if result["json"]:
                    st.caption(f"JSON 결과 {len(result['json'])}건")
                    st.json(result["json"])
                # else:
                #     st.info("JSON 결과가 없습니다. 아래 원문 로그(STDOUT/STDERR)를 확인하세요.")
                st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)

                # 디버깅용
                # if result["non_json_lines"]:
                #     with st.expander("STDOUT 원문 로그", expanded=True):
                #         st.code("\n".join(result["non_json_lines"]))
                # elif (result.get("stdout") or "").strip():
                #     with st.expander("STDOUT 원문 로그", expanded=True):
                #         st.code(result["stdout"])

                # if result["stderr"].strip():
                #     with st.expander("STDERR 로그", expanded=True):
                #         st.code(result["stderr"])
                st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)

        
        with tab2:
            st.markdown("""
            <div style="
                background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
                padding: 26px 30px;
                border-radius: 18px;
                margin-top: 18px;
                margin-bottom: 28px;
                border: 1px solid #e2e8f0;
                box-shadow: 0 4px 12px rgba(0,0,0,0.04);
            ">
                <div style="
                    display: flex;
                    align-items: center;
                    gap: 8px;
                    font-weight: 700;
                    font-size: 17px;
                    margin-bottom: 22px;
                ">
                    💻 명령어 사용 가이드
                </div>
                <div style="font-size:15px; line-height:1.8;">
                    <div style="font-weight:600; margin-bottom:4px;">
                        웹 스캔 기본 형식
                    </div>
                    <div style="color:#475569; margin-bottom:18px;">
                        nuclei -u (URL) -t (템플릿경로)
                    </div>
                    <div style="font-weight:600; margin-bottom:4px;">
                        네트워크/IP 스캔 기본 형식
                    </div>
                    <div style="color:#475569; margin-bottom:20px;">
                        nuclei -target (IP) -t (템플릿경로)
                    </div>
                    <div style="font-weight:600; margin-bottom:6px;">
                        예시
                    </div>
                    <div style="color:#475569; line-height:1.7;">
                        nuclei -u https://xxxx.com -t nuclei-templates/http/cves -severity critical,high<br>
                        nuclei -target 192.168.xxx.xxx -t nuclei-templates/network/cves
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)

            manual_cmd = st.text_input(
                "Nuclei 명령어",
                placeholder="명령어를 입력하세요",
                key="nuclei_manual_cmd",
            )
            st.caption("** 입력한 명령어를 로컬 터미널처럼 실행합니다.보안을 위해 `nuclei` 명령어만 허용됩니다.")
            st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)
            
            if st.button("🚀 nuclei 명령어 실행", use_container_width=True, key="nuclei_manual_run"):
                result, err = execute_nuclei_command(manual_cmd)
                if err:
                    st.error(err)
                else:
                    st.session_state["nuclei_last_result"] = result
                    st.success("명령어 실행이 완료되었습니다.")
            
            # ===============================
            # 📊 실행 결과 출력 (tab2 전용)
            # ===============================
            result = st.session_state.get("nuclei_last_result")
            if result:
                st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)
                st.markdown("#### 📊 실행 결과")

                st.caption(
                    f"JSON 건수: `{len(result.get('json', []))}` | "
                    f"STDOUT 라인: `{len((result.get('stdout') or '').splitlines())}` | "
                    f"STDERR 라인: `{len((result.get('stderr') or '').splitlines())}`"
                )
                st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

                no_findings = (
                    result.get("returncode", 1) == 0
                    and len(result.get("json", [])) == 0
                    and not (result.get("stdout") or "").strip()
                    and not (result.get("stderr") or "").strip()
                )
                if no_findings:
                    st.success("탐지된 취약점이 없습니다. (No findings)")

                if result.get("json"):
                    st.caption(f"JSON 결과 {len(result['json'])}건")
                    st.json(result["json"])

                stdout_text = (result.get("stdout") or "")
                stderr_text = (result.get("stderr") or "")

                with st.expander(f"📤 STDOUT 보기 ({len(stdout_text.splitlines())} lines)", expanded=True):
                    if stdout_text.strip():
                        st.code(stdout_text, language="bash")
                    else:
                        st.caption("STDOUT 출력이 없습니다.")

                with st.expander(f"📥 STDERR 보기 ({len(stderr_text.splitlines())} lines)", expanded=False):
                    if stderr_text.strip():
                        st.code(stderr_text, language="bash")
                    else:
                        st.caption("STDERR 출력이 없습니다.")

                st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)
                
        # st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)
        # st.divider()
        # st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
        st.markdown("""
            <hr style="
                border: none;
                height: 1px;
                background-color: #ddd;
                margin: 0;
                width: 100vw;
                position: relative;
                left: 50%;
                transform: translateX(-50%);
            ">
            """, unsafe_allow_html=True)
        

# =========================================================
# 기록 페이지
# =========================================================
elif st.session_state.page == "history":
    cleanup_reports()
    
    # ===============================
    # 배너
    # ===============================
    st.markdown("""
        <div style="
            width: 100%;
            overflow: hidden;
            box-shadow: 0 8px 24px rgba(0,0,0,0.08);
            margin-bottom: 32px;
        ">
            <img src="https://images.unsplash.com/photo-1550751827-4bd374c3f58b"
                style="width:100%; height:220px; object-fit:cover;">
        </div>
        """, unsafe_allow_html=True)

    
    # ===============================
    # 진단 결과
    # ===============================
    st.markdown("""
    <div class="diagnosis-wrapper">
        <div class="diagnosis-card">
            <h3 class="diagnosis-title">⚙️ 진단 결과</h3>
            <div class="diagnosis-desc">
                저장된 진단 결과를 보관합니다.
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("<div style='height:80px'></div>", unsafe_allow_html=True)

    _, center, _ = st.columns([1, 3, 1])

    with center:
        st.markdown(
            """
            <div style="
                font-size: 20px;
                font-weight: 500;
                margin-bottom: 12px;
            ">
                📂 보관함
            </div>
            """,
            unsafe_allow_html=True
        )
        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        HISTORY_DIR = CURRENT_DIR / "history"
        HISTORY_DIR.mkdir(exist_ok=True)

        files = sorted(HISTORY_DIR.glob("*.xlsx"), reverse=True)

        if not files:
            st.info("저장된 진단 기록이 없습니다.")
        else:
            for f in files:
                col_file, col_del = st.columns([8, 2])
                
                with col_file:
                    with open(f, "rb") as file_data:
                        st.download_button(
                            label=f"📄 {f.name}",
                            data=file_data,
                            file_name=f.name,
                            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                            key=f"download_{f.name}",
                            use_container_width=True
                        )
                
                with col_del:
                    if st.button("삭제", key=f"del_{f.name}", use_container_width=True):

                        import os
                        
                        if f.exists():
                            f.unlink()
                        
                        txt_filename = f.name.replace(".docx", ".txt")
                        txt_file = REPORTS_DIR / txt_filename
                        
                        if txt_file.exists():
                            txt_file.unlink()
                            st.success(f"{f.name} 및 리포트 삭제 완료")
                        else:
                            st.warning(f"파일은 삭제되었으나, {txt_filename} 파일을 찾을 수 없습니다.")
                        
                        st.rerun()

    st.markdown(
        "<div style='flex:1'></div>",
        unsafe_allow_html=True
    )

# =========================================================
# footer
# =========================================================
st.markdown(f"""
<style>
.app-footer {{
    width: 100%;
    margin-top: auto;
    margin-bottom: 0 !important;
    padding: 12px 0;
    border-top: 1px solid #e5e5e5;
    background-color: #ffffff;
}}

.footer-inner {{
    max-width: 1100px;
    margin: auto;
    display: flex;
    justify-content: center;
    align-items: center;
    gap: 48px;
}}

.footer-inner img {{
    height: 48px;
    object-fit: contain;
    opacity: 0.9;
}}
</style>

<div class="app-footer">
    <div class="footer-inner">
        <img src="data:image/png;base64,{RAPA_LOGO}" alt="RAPA">
        <img src="data:image/png;base64,{AUTOEVER_LOGO}" alt="Hyundai AutoEver">
    </div>
</div>
""", unsafe_allow_html=True)
